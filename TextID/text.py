import sys
from PyQt5 import QtCore, QtGui, QtWidgets, uic
# from PyQt5.QtWidgets import QApplication, QMessageBox, QFileDialog, QWidget, QLineEdit
# from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import QDateTime, Qt, QDate, QThread, pyqtSignal
import re
from conf import settings
import os
import openpyxl
from datetime import datetime


def IntFormat(digital):
    pattern = re.compile('.{2}')
    temp = "{:#X}".format(digital)
    temp_str = str(temp).replace('0X', '')
    id_str = '0x' + ',0x'.join(pattern.findall(temp_str))
    return id_str


def MatchKey(mlist, content):
    for n in range(len(mlist)):
        temp = content.strip()
        if temp == mlist[n]['content']:
            temp_id = mlist[n]['id']
            return temp_id
    return ''


def CheckFileName(filepath, name):
    filelist = os.listdir(filepath)
    for filename in filelist:
        if name in filename.upper():
            return filename


# 列表去重
def func(mlist):
    return list(set(mlist))


def weekday_to_chinese(weekday_abbr):
    weekday_dict = {
        'Mon': '周一',
        'Tue': '周二',
        'Wed': '周三',
        'Thu': '周四',
        'Fri': '周五',
        'Sat': '周六',
        'Sun': '周日'
    }
    return weekday_dict.get(weekday_abbr, weekday_abbr)


class QEventHandler(QtCore.QObject):
    def eventFilter(self, obj, event):
        """
        处理窗体内出现的事件，如果有需要则自行添加if判断语句；
        目前已经实现将拖到控件上文件的路径设置为控件的显示文本；
        """
        if event.type() == QtCore.QEvent.DragEnter:
            print(event.type())
            event.accept()
        if event.type() == QtCore.QEvent.Drop:
            md = event.mimeData()
            if md.hasUrls():
                # 此处md.urls()的返回值为拖入文件的file路径列表，即支持多文件同时拖入；
                # 此处默认读取第一个文件的路径进行处理，可按照个人需求进行相应的修改
                url = md.urls()[0]
                obj.setText(url.toLocalFile())
                # settings.Modify(url.toLocalFile())  # 执行写入功能后才记录修改
                return True
        return super().eventFilter(obj, event)


class ConvertThread(QThread):
    update_date = pyqtSignal(float, float, float, str, str, str, str)

    def __init__(self, input_text, start_id_text, start_id_cmd, start_id_exp, checkbox_11, set_path):
        super(ConvertThread, self).__init__()
        self.input_text = input_text
        self.start_id_text = start_id_text
        self.start_id_cmd = start_id_cmd
        self.start_id_exp = start_id_exp
        self.checkbox_11 = checkbox_11
        self.set_path = set_path

    def GetTextName(self):
        txt_list = os.listdir(self.set_path)
        for txt in txt_list:
            if re.fullmatch('.txt', os.path.splitext(txt)[1], re.I) and re.match('text', os.path.splitext(txt)[0],
                                                                                 re.I):
                return txt

    def Find(self, search_content):
        path = self.set_path + '/' + self.GetTextName()
        f = open(path, 'r', encoding='gbk')
        while True:
            line_text = f.readline()
            if not line_text:  # 文件末尾
                break
            if not line_text.strip():  # 空白行
                continue
            if line_text[0] == ';':  # 注释行
                continue

            start_pos = line_text.find('\"')
            end_pos = line_text.rfind('\"')
            text_list = [line_text[0:start_pos].strip(), line_text[start_pos + 1:end_pos].strip()]

            # text_list = line_text.split('\"')  # text
            text_part = text_list[1]
            if search_content == text_part:
                id_text = text_list[0]
                if id_text[-1] == ',':
                    id_text = id_text[:-1]
                f.close()
                return line_text, id_text
        f.close()
        return '', ''

    @staticmethod
    def add_new_id(id_, text):
        lineWidth = 60
        id_str = IntFormat(id_)
        temp = id_str + ',\t' + '\"' + text + '\"'
        # size = 65-len(temp.encode('gbk'))
        # print(100 - len(temp.encode('gbk')), sys.getsizeof(temp))
        # print(temp)
        # temp1 = temp.ljust(size, ' ')
        if len(temp.encode('gbk')) < lineWidth:
            temp += ' ' * (lineWidth - len(temp.encode('gbk')))
        id_note = '    ;;[{}][][]\n'.format(hex(id_))
        new_id = temp + id_note
        # print(new_id)
        return new_id

    def run(self):
        text_id = int(self.start_id_text, 16) & 0xFFFFFFFF
        cmd_id = int(self.start_id_cmd, 16) & 0xFFFFFFFF
        exp_id = int(self.start_id_exp, 16) & 0xFFFFFFFF
        text_list = self.input_text.split('\n')
        new_text_list = []
        cmd_list = []
        exp_list = []
        old_text = ''
        new_text = ''
        cmd_text = ''
        exp_text = ''
        for text in text_list:
            exp_rule_01 = False
            exp_rule_02 = False
            text = text.strip()
            if text != '':
                regex1 = re.compile(r'^[\da-fA-F ]+($|;)')
                result = regex1.match(text)  # 匹配命令文本

                # 以#开始，以#结束，#标定成功# 也被当成算法 y=string(标定成功);
                regex2 = re.compile('(if\(.*)|(strcat:.*)|(y=.*)|(y =.*)|(y = .*)|^(#.*#)$')
                result2 = regex2.match(text)  # 匹配算法文本
                if result:  # 命令文本
                    id_str = IntFormat(cmd_id)
                    result = result.group()
                    result = re.sub(r'[\s]+', '', result).replace(';', '')
                    len_str = str(int(len(result) / 2))
                    pattern = re.compile('.{2}')
                    cmd_str = len_str + ',0x' + ',0x'.join(pattern.findall(result))

                    if cmd_str not in cmd_list:
                        cmd_text += id_str + ',\t' + cmd_str + '\n'
                        cmd_list.append(cmd_str)
                        cmd_id += 1
                elif result2:  # 算法文本 | TEXT文本
                    id_str = IntFormat(exp_id)
                    result2 = result2.group()

                    # Ans.ByteXX替换成x
                    pos_list = re.findall(r'Ans.Byte(\d+)', result2)
                    # print(pos_list)
                    if len(pos_list) != 0:
                        pos_list = func(pos_list)
                        pos_min = min(map(int, pos_list))
                        for i in range(len(pos_list)):
                            pos = int(pos_list[i]) - pos_min
                            old = 'Ans.Byte{:d}'.format(int(pos_list[i]))
                            new = 'x{:}'.format(pos + 1)
                            result2 = result2.replace(old, new)

                    # 文本替换成y=string();\s\W.+?\s
                    # exp_text_list = re.findall(r'\s[^A-Za-z0-9_(].+?\s*' , result2)
                    # exp_text_list = re.findall(r'\[.+?\]', result2)  # if(Ans.Byte8==0) [完成] else [未完成] # 废弃带[%d]的算法不行
                    exp_text_list = re.findall(r'“.+?”', result2)  # if(Ans.Byte8==0) “完成” else “未完成”
                    if len(exp_text_list) == 0:
                        exp_text_list = re.findall(r'#.+?#', result2)  # #完成# => y=string(0x12,0x23,0x34,0x56);
                        if len(exp_text_list) != 0:
                            exp_rule_01 = True
                    if len(exp_text_list) == 0:  # 2023-12-21 add
                        exp_text_list = re.findall(r'string\((.+?)\)',
                                                   result2)  # y=string(完成) => y=string(0x12,0x23,0x34,0x56);
                        if len(exp_text_list) != 0:
                            exp_rule_02 = True
                    if len(exp_text_list) == 0:
                        temp_exp_text = result2.split('else')  # if(x1==0x00) 数据1 else if(x1==0x01) 数据2 else 无效数据
                        if len(temp_exp_text) == 0:
                            temp_exp_text = result2.split(
                                'if')  # strcat:if(Ans.Byte9.bit0==1) 速度过慢if(x1.bit1==1) 速度过快if(x1.bit2==1) 横摆角过大if(x1.bit3==1) 加速过快if(x1.bit4==1) 目标不充分if(x1.bit5==1) 雷达失明if(x1==0) 正常"
                        regex3 = re.compile(r'\) (.*?)$')
                        for temp in temp_exp_text:
                            # 纯算法部分 if(x1<20) y=SPRINTF([%d],x1);else 无效数据
                            if re.search(r'SPRINTF', temp) or re.search(r'y=', temp):
                                continue
                            if 'if' not in temp:
                                exp_text_list.append(temp.strip())
                            else:
                                temp_list = re.findall(regex3, temp)
                                if len(temp_list):
                                    exp_text_list.append(temp_list[0].strip())

                    if len(exp_text_list) != 0:
                        for m in range(len(exp_text_list)):
                            temp_text = exp_text_list[m].replace('“', '').replace('”', '')
                            if exp_rule_01:
                                temp_text = exp_text_list[m].replace('#', '')
                            if exp_rule_02:  # 2023-12-21 add
                                temp_text = exp_text_list[m].replace('(', '').replace(')', '')

                            first_old_text = ''
                            old_id = ''
                            if self.checkbox_11:
                                first_old_text, old_id = self.Find(temp_text)  # 第一个完全匹配的文本，找到就不找
                            if first_old_text:
                                if not MatchKey(new_text_list, temp_text):  # 在TEXT_XX中找到且之前没有出现过
                                    text_dic = {
                                        'id': old_id,
                                        'content': temp_text
                                    }
                                    new_text_list.append(text_dic)
                                    old_text += first_old_text
                                temp_id = old_id
                            else:
                                temp_id = MatchKey(new_text_list, temp_text)
                            if temp_id != '':
                                old = exp_text_list[m].strip()
                                new = 'y=string(' + temp_id + ');'
                                result2 = result2.replace(old, new)
                                result2 = result2.replace('提示', '')
                            else:
                                # 新增ID文本
                                add_text = exp_text_list[m].strip()
                                add_text = add_text.replace('“', '').replace('”', '')
                                if exp_rule_01:
                                    add_text = add_text.replace('#', '')
                                if exp_rule_02:  # 2023-12-21 add
                                    add_text = add_text.replace('(', '').replace(')', '')
                                if MatchKey(new_text_list, add_text) == '':
                                    id_str = IntFormat(text_id)
                                    text_dic = {
                                        'id': id_str,
                                        'content': add_text
                                    }
                                    new_text_list.append(text_dic)
                                    new_text += self.add_new_id(text_id, add_text)
                                    # text_str = '\"' + add_text + '\"'
                                    # new_text += id_str + ',\t' + text_str + '\n'
                                    text_id += 1
                                # 算法文本替换成ID格式
                                old = exp_text_list[m].strip()
                                if exp_rule_02:  # 2023-12-21 add
                                    new = id_str
                                else:
                                    new = 'y=string(' + id_str + ');'
                                result2 = result2.replace(old, new)
                                result2 = result2.replace('提示', '')

                    if result2[-1] != ';':
                        exp_str = '\"' + result2 + ';\"'
                    else:
                        exp_str = '\"' + result2 + '\"'
                    if exp_str not in exp_list:
                        id_str = IntFormat(exp_id)
                        exp_text += id_str + ',\t' + exp_str + '\n'
                        exp_list.append(exp_str)
                        exp_id += 1
                else:  # TEXT文本
                    first_old_text = ''
                    old_id = ''
                    if self.checkbox_11:
                        first_old_text, old_id = self.Find(text)  # 第一个完全匹配的文本，找到就不找
                    if first_old_text:
                        if not MatchKey(new_text_list, text):  # 在TEXT_XX中找到且之前没有出现过
                            text_dic = {
                                'id': old_id,
                                'content': text
                            }
                            new_text_list.append(text_dic)
                            old_text += first_old_text
                    else:
                        if MatchKey(new_text_list, text) == '':
                            id_str = IntFormat(text_id)
                            text_dic = {
                                'id': id_str,
                                'content': text
                            }
                            new_text_list.append(text_dic)
                            # text_str = '\"' + text + '\"'
                            # new_text += id_str + ',\t' + text_str + '\n'
                            new_text += self.add_new_id(text_id, text)
                            text_id += 1

        self.update_date.emit(text_id, cmd_id, exp_id, old_text, new_text, cmd_text, exp_text)


class AddID:
    set_path = ''
    id_dic = {}
    MAX_LINE = 500
    line = 0
    num = 0
    first = True
    text_id_str = ''
    cmd_id_str = ''
    exp_id_str = ''

    def __init__(self):
        super(AddID, self).__init__()
        self.convert_thread = None
        self.actionC = None
        self.actionB = None
        self.actionA = None
        self.model = None
        self.count = 0
        self.ui = uic.loadUi('ui/convert.ui')
        # self.ui.pushButton.clicked.connect(self.Convert)
        self.ui.pushButton.clicked.connect(self.ConvertThread)
        # self.ui.lineEdit.textChanged.connect(self.Upper)
        # pos = self.ui.lineEdit.cursorPosition()
        # self.ui.lineEdit.setCursorPosition(pos)
        self.ui.lineEdit.setMaxLength(10)

        self.ui.tabWidget.currentChanged.connect(self.on_tab_changed)

        self.ui.pushButton_2.clicked.connect(self.Select)
        self.ui.pushButton_12.clicked.connect(self.SelectExcel)
        self.ui.pushButton_3.clicked.connect(lambda: self.Search(1))  # 文本查找ID
        self.ui.pushButton_4.clicked.connect(lambda: self.Search(0))  # ID查找文本
        self.ui.pushButton_5.clicked.connect(self.CopyAnyData)
        self.ui.pushButton_6.clicked.connect(self.CopyIdData)
        self.ui.pushButton_13.clicked.connect(self.ReadExcelGetText)

        self.ui.copy_text.clicked.connect(self.Copy1)
        self.ui.copy_cmd.clicked.connect(self.Copy2)
        self.ui.copy_exp.clicked.connect(self.Copy3)

        self.ui.add_text.clicked.connect(self.Add1)
        self.ui.add_cmd.clicked.connect(self.Add2)
        self.ui.add_exp.clicked.connect(self.Add3)

        self.ui.add_all.clicked.connect(self.AddAll)

        self.ui.clearButton.clicked.connect(self.Clear)
        self.ui.clearLineButton.clicked.connect(self.RemoveLine)
        self.ui.getCmmdButton.clicked.connect(self.GetSendCommand)
        self.ui.addLine.clicked.connect(self.replace_line_feed)
        self.ui.FormattingID.clicked.connect(self.id_format_show)

        self.set_path = settings.Read()
        if self.set_path:
            self.ui.lineEdit_2.setText(self.set_path)

        excel_path = settings.ReadExcelPath()
        if excel_path:
            self.ui.lineEdit_11.setText(excel_path)

        self.id_dic = settings.ReadID()
        if self.id_dic.get('text_id'):
            self.ui.lineEdit.setText(self.id_dic.get('text_id'))
        if self.id_dic.get('command_id'):
            self.ui.lineEdit_5.setText(self.id_dic.get('command_id'))
        if self.id_dic.get('express_id'):
            self.ui.lineEdit_6.setText(self.id_dic.get('express_id'))

        self.ui.lineEdit_2.installEventFilter(QEventHandler(self.ui))
        self.ui.lineEdit_11.installEventFilter(QEventHandler(self.ui))

        self.InitTableView()

        # 右键菜单，如果不设为CustomContextMenu,无法使用customContextMenuRequested
        self.ui.tableView.setContextMenuPolicy(Qt.CustomContextMenu)
        self.ui.tableView.customContextMenuRequested.connect(self.showContextMenu)

        if not self.RefreshTxtList():
            self.ui.checkBox.setChecked(False)
        self.ui.checkBox.stateChanged.connect(self.RefreshTxtList)
        self.ui.checkBox_2.stateChanged.connect(self.RefreshTxtList)

        if settings.ReadDuplicateRemoval() == 'True':
            self.ui.checkBox_11.setChecked(True)
        else:
            self.ui.checkBox_11.setChecked(False)
        self.ui.checkBox_11.stateChanged.connect(self.DuplicateRemovalState)
        self.ui.lineEdit_4.setText(f'{self.show_date()}\t吴琛')
        # self.ui.lineEdit_4.setText('Current date: ' + QDate.currentDate().toString() + '\tDeveloper: 吴琛')
        # self.ui.status = self.statusBar()
        # self.statusBar().showMessage("当前时间：" + QDateTime.currentDateTime().toString(Qt.DefaultLocaleLongDate), 0)
        # self.show_1 = QLabel("作者：WUCHEN")
        # self.show_2 = QLabel("邮箱：1036542915@qq.com")
        # self.ui.status.addPermanentWidget(self.show_1, stretch=0)  # 比例
        # self.ui.status.addPermanentWidget(self.show_2, stretch=0)

    @staticmethod
    def show_date():
        now = datetime.now()
        date_time_str = now.strftime('%Y年%m月%d日 %H时%M分%S秒')
        weekday_abbr = now.strftime('%a')
        chinese_weekday = ' ' + weekday_to_chinese(weekday_abbr)
        return date_time_str + chinese_weekday

    def on_tab_changed(self):
        # print(f"Tab changed from {self.ui.tabWidget.currentIndex()}")
        self.ui.lineEdit_4.setText(f'{self.show_date()}\t吴琛')

    def ButtonEnabled(self, state):
        self.ui.pushButton_13.setEnabled(state)
        self.ui.pushButton_12.setEnabled(state)
        self.ui.add_all.setEnabled(state)
        self.ui.add_text.setEnabled(state)
        self.ui.add_cmd.setEnabled(state)
        self.ui.add_exp.setEnabled(state)
        self.ui.pushButton.setEnabled(state)
        self.ui.copy_text.setEnabled(state)
        self.ui.copy_cmd.setEnabled(state)
        self.ui.copy_exp.setEnabled(state)
        self.ui.getCmmdButton.setEnabled(state)
        self.ui.clearLineButton.setEnabled(state)
        self.ui.clearButton.setEnabled(state)

    def GetTextName(self):
        txt_list = os.listdir(self.set_path)
        for txt in txt_list:
            if re.fullmatch('.txt', os.path.splitext(txt)[1], re.I) and re.match('text', os.path.splitext(txt)[0],
                                                                                 re.I):
                return txt

    def Find(self, search_content):
        path = self.set_path + '/' + self.GetTextName()
        f = open(path, 'r', encoding='gbk')
        while True:
            line_text = f.readline()
            if not line_text:  # 文件末尾
                break
            if not line_text.strip():  # 空白行
                continue
            if line_text[0] == ';':  # 注释行
                continue

            text_list = line_text.split('\"')  # text
            text_part = text_list[1]
            if re.fullmatch(search_content, text_part):
                id_text = text_list[0].strip()
                if id_text[-1] == ',':
                    id_text = id_text[:-1]
                f.close()
                return line_text, id_text
        f.close()
        return '', ''

    def DuplicateRemovalState(self):
        state = self.ui.checkBox_11.isChecked()
        if state:
            settings.ModifyDuplicateRemoval('True')
        else:
            settings.ModifyDuplicateRemoval('False')

    def parseData(self, cmd_text):
        cmd_output_list = []
        multi_frame = 0
        all_cmd_bytes = []
        cmd_len = 0
        if not re.findall(r'req.*', cmd_text.strip(), re.I):  # 不存在req开头的不是命令
            # text_list = cmd_text.split(':')
            # text_list = re.split('[:：]', cmd_text.replace('\n', ''))
            text_lists = cmd_text.strip().split('\n')
            regexp = re.compile('if.*|strcat:.*|y=.*|y =.*|y = .*|else.*')
            exp_text = ''
            for text in text_lists:
                flag = regexp.findall(text.strip())  # 匹配算法文本
                if not flag:
                    if exp_text:  # 不是算法，且前面一个算法结束
                        cmd_output_list.append(exp_text)
                        exp_text = ''
                    cmd_output_list.append(text.strip())
                else:
                    exp_text += text.strip()
            if exp_text:
                cmd_output_list.append(exp_text)
            return cmd_output_list
        cmd_lists = cmd_text.strip().split('\n')
        req_cmd_lists = [_ for _ in cmd_lists if not re.findall(r'ans', _, re.I) and _]  # 去掉应答命令
        # print(req_cmd_lists)
        multi_cmd_flag = False
        head_cmd_flag = False
        for cmd in req_cmd_lists:
            cmd = cmd.strip()
            temp = re.findall(r'Req:.*?([\da-fA-F ]+).*?', cmd, re.I)
            if not temp:  # 不是Req开头的显示原文
                cmd_output_list.append(cmd)
                continue
            cmd_bytes = temp[0].strip().split(' ')
            # print(cmd_bytes)
            if cmd_bytes[0] == '08' and len(cmd_bytes) == 11:  # 标准CAN
                if cmd_bytes[3][0] == '1':
                    multi_cmd_flag = True
                    head_cmd_flag = True
                    multi_frame = 0
                    cmd_len = 0
                    all_cmd_bytes = []
                    del cmd_bytes[3]

                    if head_cmd_flag:
                        cmd_len = int(cmd_bytes[3], 16)
                        # print(cmd_len)
                        cmd_len -= 6
                        if cmd_len % 7 > 0:
                            multi_frame = (cmd_len // 7) + 1
                        else:
                            multi_frame = cmd_len // 7
                        all_cmd_bytes = cmd_bytes
                        head_cmd_flag = False
                        continue
                if not multi_cmd_flag and not head_cmd_flag:
                    cmd_output_list.append(temp[0].strip())  # 单帧

                if multi_cmd_flag and not head_cmd_flag:
                    if int(cmd_bytes[3], 16) < 20:
                        cmd_output_list.append(' '.join(all_cmd_bytes))
                        multi_cmd_flag = False
                        self.ui.lineEdit_4.setText('存在异常多帧数据')
                        cmd_output_list.append(temp[0].strip())  # 单帧

                if multi_cmd_flag and not head_cmd_flag:
                    if multi_frame == 1:
                        if cmd_len % 7:
                            all_cmd_bytes += cmd_bytes[4:4 + (cmd_len % 7)]
                        else:
                            all_cmd_bytes += cmd_bytes[4:]
                        cmd_output_list.append(' '.join(all_cmd_bytes))
                        multi_cmd_flag = False
                    else:
                        all_cmd_bytes += cmd_bytes[4:]
                    if len(req_cmd_lists) < multi_frame + 1:  # 只有一条多帧命令且帧数不够
                        cmd_output_list.append(' '.join(all_cmd_bytes))
                        multi_cmd_flag = False
                        self.ui.lineEdit_4.setText('存在异常多帧数据')
                    multi_frame -= 1
            elif cmd_bytes[0] == '88' and len(cmd_bytes) == 13:  # 扩展CAN
                if cmd_bytes[3 + 2][0] == '1':
                    multi_cmd_flag = True
                    head_cmd_flag = True
                    multi_frame = 0
                    cmd_len = 0
                    all_cmd_bytes = []
                    del cmd_bytes[3 + 2]

                    if head_cmd_flag:
                        cmd_len = int(cmd_bytes[3 + 2], 16)
                        # print(cmd_len)
                        cmd_len -= 6
                        if cmd_len % 7 > 0:
                            multi_frame = (cmd_len // 7) + 1
                        else:
                            multi_frame = cmd_len // 7
                        all_cmd_bytes = cmd_bytes
                        head_cmd_flag = False
                        continue
                if not multi_cmd_flag and not head_cmd_flag:
                    cmd_output_list.append(temp[0].strip())  # 单帧

                if multi_cmd_flag and not head_cmd_flag:
                    if int(cmd_bytes[5], 16) < 20:
                        cmd_output_list.append(' '.join(all_cmd_bytes))
                        multi_cmd_flag = False
                        self.ui.lineEdit_4.setText('存在异常多帧数据')
                        cmd_output_list.append(temp[0].strip())  # 单帧

                if multi_cmd_flag and not head_cmd_flag:
                    if multi_frame == 1:
                        if cmd_len % 7:
                            all_cmd_bytes += cmd_bytes[4 + 2:4 + 2 + (cmd_len % 7)]
                        else:
                            all_cmd_bytes += cmd_bytes[4 + 2:]
                        cmd_output_list.append(' '.join(all_cmd_bytes))
                        multi_cmd_flag = False
                    else:
                        all_cmd_bytes += cmd_bytes[4 + 2:]
                    if len(req_cmd_lists) < multi_frame + 1:  # 只有一条多帧命令且帧数不够
                        cmd_output_list.append(' '.join(all_cmd_bytes))
                        multi_cmd_flag = False
                        self.ui.lineEdit_4.setText('存在异常多帧数据')
                    multi_frame -= 1
            else:
                cmd_output_list.append(temp[0].strip())
        return cmd_output_list

    def GetSendCommand(self):
        # 只从剪切板取值
        clipboard = QApplication.clipboard()
        cmd_text = clipboard.text()
        # print(cmd_text)
        cmd_output_list = self.parseData(cmd_text)

        # 复制到剪切板
        if len(cmd_output_list) >= 2:
            data = '\n'.join(cmd_output_list)
        else:
            data = ''.join(cmd_output_list).strip()
        self.ui.lineEdit_4.setText('提取发送命令完成')
        clipboard = QApplication.clipboard()
        clipboard.setText(data)

        # 显示在plainTextEdit控件上
        input_text = self.ui.plainTextEdit.toPlainText()
        if input_text:
            if input_text[-1] != '\n':
                input_text += '\n'
        input_text += data + '\n'
        self.ui.plainTextEdit.setPlainText(input_text)

    @staticmethod
    def RemoveLine():
        # 只从剪切板取值
        clipboard = QApplication.clipboard()
        tag_str = clipboard.text()

        if tag_str:
            clipboard.setText(tag_str.replace('\n', ''))

    def ClearTable(self):
        self.InitTableView()

    def showContextMenu(self):  # 创建右键菜单
        self.ui.tableView.contextMenu = QMenu()
        self.actionA = self.ui.tableView.contextMenu.addAction('复制ID')
        self.actionB = self.ui.tableView.contextMenu.addAction('复制原文本')
        self.actionC = self.ui.tableView.contextMenu.addAction('清空')
        # self.actionA = self.view.contextMenu.exec_(self.mapToGlobal(pos))  # 1
        self.ui.tableView.contextMenu.popup(QCursor.pos())
        self.actionA.triggered.connect(self.CopyIdData)
        self.actionB.triggered.connect(self.CopyAnyData)
        self.actionC.triggered.connect(self.ClearTable)
        # self.view.contextMenu.move(self.pos())  # 3
        self.ui.tableView.contextMenu.show()

    def InitTableView(self):
        # 设置表格的行列数
        self.model = QStandardItemModel(self.MAX_LINE, 3)

        # self.model = QStandardItemModel(self.MAX_LINE)
        # self.model.setRowCount(self.MAX_LINE)
        # self.model.setColumnCount(3)

        # # 获取表头的引用
        # header = self.ui.tableView.horizontalHeader()
        #
        # # 设置第一列和第三列的宽度为固定
        # header.setSectionResizeMode(0, QHeaderView.Fixed)
        # header.resizeSection(0, 100)  # 设置第一列固定宽度为100
        # header.setSectionResizeMode(2, QHeaderView.Fixed)
        # header.resizeSection(2, 150)  # 设置第三列固定宽度为150
        #
        # # 设置中间列（索引为1的列）自动拉伸
        # header.setSectionResizeMode(1, QHeaderView.Stretch)

        # 设置表头
        self.model.setHorizontalHeaderLabels(['文本ID', '文本内容', '文件名'])

        # 设置表头背景色
        self.ui.tableView.horizontalHeader().setStyleSheet(
            "QHeaderView::section{background-color: rgb(170, 255, 127);}")

        # 让滚动条滚动到最顶部
        self.ui.tableView.scrollToTop()

        # self.ui.tableView.resizeColumnsToContents()
        # self.ui.tableView.resizeRowsToContents()

        table_widget = self.ui.tableView
        table_width = table_widget.width()
        col_1_width = table_width // 4
        col_2_width = table_width // 2
        col_3_width = table_width // 4
        # print(table_widget.size(), table_width, col_1_width, col_2_width, col_3_width)
        table_widget.setColumnWidth(0, col_1_width)
        table_widget.setColumnWidth(1, col_2_width)
        table_widget.setColumnWidth(2, col_3_width)

        # table_widget.setColumnWidth(0, 86)
        # table_widget.setColumnWidth(1, 683)
        # table_widget.setColumnWidth(2, 85)
        table_widget.setWordWrap(True)

        # 设置单元格的背脊颜色为红
        # table_widget.setForeground(QBrush(QColor(255, 0, 0)))

        # self.ui.tableView.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        # self.ui.tableView.horizontalHeader().resizeSection(0, 100)
        # self.ui.tableView.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        # self.ui.tableView.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        # self.ui.tableView.horizontalHeader().resizeSection(2, 200)
        # self.ui.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.ui.tableView.resizeColumnToContents(1)
        # self.ui.tableView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        # 水平方向标签拓展剩下的窗口部分，填满表格
        # self.ui.tableView.horizontalHeader().setStretchLastSection(True)
        # 水平方向，表格大小拓展到适当的尺寸
        # self.ui.tableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.tableView.setModel(self.model)

    def Clear(self):
        self.ui.plainTextEdit.clear()
        self.ui.textEdit_8.clear()
        self.ui.textEdit_2.clear()
        self.ui.textEdit_3.clear()
        self.ui.textEdit_4.clear()

    def Copy1(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.ui.textEdit_2.toPlainText())

    def Copy2(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.ui.textEdit_3.toPlainText())

    def Copy3(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.ui.textEdit_4.toPlainText())

    def Upper(self):
        text = self.ui.lineEdit.text()
        upp_text = text.upper()
        self.ui.lineEdit.setText(upp_text)

    '''def Convert(self):
        start_id_text = self.ui.lineEdit.text().strip()
        if len(start_id_text) == 0:
            QMessageBox.warning(self.ui, '错误信息', '最少输入1位16进制TEXT起始ID')
            return False

        start_id_cmd = self.ui.lineEdit_5.text().strip()
        if len(start_id_cmd) == 0:
            QMessageBox.warning(self.ui, '错误信息', '最少输入1位16进制COMMAND起始ID')
            return False

        start_id_exp = self.ui.lineEdit_6.text().strip()
        if len(start_id_exp) == 0:
            QMessageBox.warning(self.ui, '错误信息', '最少输入1位16进制EXPRESS起始ID')
            return False

        check_re = re.compile(r'^[\da-fA-F]{1,8}$')
        start_id_text = start_id_text.replace('0x', '').replace('0X', '')
        if not check_re.fullmatch(start_id_text):
            QMessageBox.warning(self.ui, 'TEXT', '输入字符包含非16进制数或者长度大于8！')
            return False

        start_id_cmd = start_id_cmd.replace('0x', '').replace('0X', '')
        if not check_re.fullmatch(start_id_cmd):
            QMessageBox.warning(self.ui, 'COMMAND', '输入字符包含非16进制数或者长度大于8！')
            return False

        start_id_exp = start_id_exp.replace('0x', '').replace('0X', '')
        if not check_re.fullmatch(start_id_exp):
            QMessageBox.warning(self.ui, 'EXPRESS', '输入字符包含非16进制数或者长度大于8！')
            return False

        start_id_text = start_id_text.ljust(8, '0')
        text_id = int(start_id_text, 16)
        self.id_dic['text_id'] = start_id_text

        start_id_cmd = start_id_cmd.ljust(8, '0')
        cmd_id = int(start_id_cmd, 16)
        self.id_dic['command_id'] = start_id_cmd

        start_id_exp = start_id_exp.ljust(8, '0')
        exp_id = int(start_id_exp, 16)
        self.id_dic['express_id'] = start_id_exp

        # text_id = cmd_id = exp_id = start_id

        input_text = self.ui.plainTextEdit.toPlainText()
        if len(input_text) == 0:
            QMessageBox.about(self.ui, '错误信息', '文本编辑区为空')
            return False
        text_list = input_text.split('\n')
        new_text_list = []
        cmd_list = []
        exp_list = []
        old_text = ''
        new_text = ''
        cmd_text = ''
        exp_text = ''
        for text in text_list:
            exp_rule_01 = False
            text = text.strip()
            if text != '':
                regex1 = re.compile(r'^[\da-fA-F ]+($|;)')
                result = regex1.match(text)  # 匹配命令文本

                # 以#开始，以#结束，#标定成功# 也被当成算法 y=string(标定成功);
                regex2 = re.compile('(if\(.*)|(strcat:.*)|(y=.*)|(y =.*)|(y = .*)|^(#.*#)$')
                result2 = regex2.match(text)  # 匹配算法文本
                if result:  # 命令文本
                    id_str = IntFormat(cmd_id)
                    result = result.group()
                    result = re.sub(r'[\s]+', '', result).replace(';', '')
                    len_str = str(int(len(result) / 2))
                    pattern = re.compile('.{2}')
                    cmd_str = len_str + ',0x' + ',0x'.join(pattern.findall(result))

                    if cmd_str not in cmd_list:
                        cmd_text += id_str + ',\t' + cmd_str + '\n'
                        cmd_list.append(cmd_str)
                        cmd_id += 1
                elif result2:  # 算法文本 | TEXT文本
                    id_str = IntFormat(exp_id)
                    result2 = result2.group()

                    # Ans.ByteXX替换成x
                    pos_list = re.findall(r'Ans.Byte(\d+)', result2)
                    # print(pos_list)
                    if len(pos_list) != 0:
                        pos_list = func(pos_list)
                        pos_min = min(map(int, pos_list))
                        for i in range(len(pos_list)):
                            pos = int(pos_list[i]) - pos_min
                            old = 'Ans.Byte{:d}'.format(int(pos_list[i]))
                            new = 'x{:}'.format(pos + 1)
                            result2 = result2.replace(old, new)

                    # 文本替换成y=string();\s\W.+?\s
                    # exp_text_list = re.findall(r'\s[^A-Za-z0-9_(].+?\s*' , result2)
                    # exp_text_list = re.findall(r'\[.+?\]', result2)  # if(Ans.Byte8==0) [完成] else [未完成] # 废弃带[%d]的算法不行
                    exp_text_list = re.findall(r'“.+?”', result2)  # if(Ans.Byte8==0) “完成” else “未完成”
                    if len(exp_text_list) == 0:
                        exp_text_list = re.findall(r'#.+?#', result2)  # #完成# => y=string(0x12,0x23,0x34,0x56);
                        exp_rule_01 = True
                    if len(exp_text_list) == 0:
                        temp_exp_text = result2.split('else')  # if(x1==0x00) 数据1 else if(x1==0x01) 数据2 else 无效数据
                        regex3 = re.compile(r'\) (.*?)$')
                        for temp in temp_exp_text:
                            # 纯算法部分 if(x1<20) y=SPRINTF([%d],x1);else 无效数据
                            if re.search(r'SPRINTF', temp) or re.search(r'y=', temp):
                                continue
                            if 'if' not in temp:
                                exp_text_list.append(temp.strip())
                            else:
                                temp_list = re.findall(regex3, temp)
                                if len(temp_list):
                                    exp_text_list.append(temp_list[0].strip())

                    if len(exp_text_list) != 0:
                        for m in range(len(exp_text_list)):
                            temp_text = exp_text_list[m].replace('“', '').replace('”', '')
                            if exp_rule_01:
                                temp_text = exp_text_list[m].replace('#', '')

                            first_old_text = ''
                            old_id = ''
                            if self.ui.checkBox_11.isChecked():
                                first_old_text, old_id = self.Find(temp_text)  # 第一个完全匹配的文本，找到就不找
                            if first_old_text:
                                if not MatchKey(new_text_list, temp_text):  # 在TEXT_XX中找到且之前没有出现过
                                    text_dic = {
                                        'id': old_id,
                                        'content': temp_text
                                    }
                                    new_text_list.append(text_dic)
                                    old_text += first_old_text
                                temp_id = old_id
                            else:
                                temp_id = MatchKey(new_text_list, temp_text)
                            if temp_id != '':
                                old = exp_text_list[m].strip()
                                new = 'y=string(' + temp_id + ');'
                                result2 = result2.replace(old, new)
                                result2 = result2.replace('提示', '')
                            else:
                                # 新增ID文本
                                add_text = exp_text_list[m].strip()
                                add_text = add_text.replace('“', '').replace('”', '')
                                if exp_rule_01:
                                    add_text = add_text.replace('#', '')
                                if MatchKey(new_text_list, add_text) == '':
                                    id_str = IntFormat(text_id)
                                    text_dic = {
                                        'id': id_str,
                                        'content': add_text
                                    }
                                    new_text_list.append(text_dic)
                                    text_str = '\"' + add_text + '\"'
                                    new_text += id_str + ',\t' + text_str + '\n'
                                    text_id += 1
                                # 算法文本替换成ID格式
                                old = exp_text_list[m].strip()
                                new = 'y=string(' + id_str + ');'
                                result2 = result2.replace(old, new)
                                result2 = result2.replace('提示', '')

                    if result2[-1] != ';':
                        exp_str = '\"' + result2 + ';\"'
                    else:
                        exp_str = '\"' + result2 + '\"'
                    if exp_str not in exp_list:
                        id_str = IntFormat(exp_id)
                        exp_text += id_str + ',\t' + exp_str + '\n'
                        exp_list.append(exp_str)
                        exp_id += 1
                else:  # TEXT文本
                    first_old_text = ''
                    old_id = ''
                    if self.ui.checkBox_11.isChecked():
                        first_old_text, old_id = self.Find(text)  # 第一个完全匹配的文本，找到就不找
                    if first_old_text:
                        if not MatchKey(new_text_list, text):  # 在TEXT_XX中找到且之前没有出现过
                            text_dic = {
                                'id': old_id,
                                'content': text
                            }
                            new_text_list.append(text_dic)
                            old_text += first_old_text
                    else:
                        if MatchKey(new_text_list, text) == '':
                            id_str = IntFormat(text_id)
                            text_dic = {
                                'id': id_str,
                                'content': text
                            }
                            new_text_list.append(text_dic)
                            text_str = '\"' + text + '\"'
                            new_text += id_str + ',\t' + text_str + '\n'
                            text_id += 1

        # print(hex(text_id))
        self.text_id_str = hex(text_id).replace('0x', '').upper()
        self.cmd_id_str = hex(cmd_id).replace('0x', '').upper()
        self.exp_id_str = hex(exp_id).replace('0x', '').upper()
        self.ui.textEdit_8.setText(old_text)
        self.ui.textEdit_2.setText(new_text)
        self.ui.textEdit_3.setText(cmd_text)
        self.ui.textEdit_4.setText(exp_text)
        return True'''

    def GetData(self):
        start_id_text = self.ui.lineEdit.text().strip()
        if len(start_id_text) == 0:
            QMessageBox.warning(self.ui, '错误信息', '最少输入1位16进制TEXT起始ID')
            return '', start_id_text, '', ''

        start_id_cmd = self.ui.lineEdit_5.text().strip()
        if len(start_id_cmd) == 0:
            QMessageBox.warning(self.ui, '错误信息', '最少输入1位16进制COMMAND起始ID')
            return '', start_id_text, start_id_cmd, ''

        start_id_exp = self.ui.lineEdit_6.text().strip()
        if len(start_id_exp) == 0:
            QMessageBox.warning(self.ui, '错误信息', '最少输入1位16进制EXPRESS起始ID')
            return '', start_id_text, start_id_cmd, start_id_exp

        check_re = re.compile(r'^[\da-fA-F]{1,8}$')
        start_id_text = start_id_text.replace('0x', '').replace('0X', '')
        if not check_re.fullmatch(start_id_text):
            QMessageBox.warning(self.ui, 'TEXT', '输入字符包含非16进制数或者长度大于8！')
            return None

        start_id_cmd = start_id_cmd.replace('0x', '').replace('0X', '')
        if not check_re.fullmatch(start_id_cmd):
            QMessageBox.warning(self.ui, 'COMMAND', '输入字符包含非16进制数或者长度大于8！')
            return None

        start_id_exp = start_id_exp.replace('0x', '').replace('0X', '')
        if not check_re.fullmatch(start_id_exp):
            QMessageBox.warning(self.ui, 'EXPRESS', '输入字符包含非16进制数或者长度大于8！')
            return None

        start_id_text = start_id_text.ljust(8, '0')
        self.id_dic['text_id'] = start_id_text

        start_id_cmd = start_id_cmd.ljust(8, '0')
        self.id_dic['command_id'] = start_id_cmd

        start_id_exp = start_id_exp.ljust(8, '0')
        self.id_dic['express_id'] = start_id_exp

        input_text = self.ui.plainTextEdit.toPlainText()
        if len(input_text) == 0:
            QMessageBox.about(self.ui, '错误信息', '文本编辑区为空')
            return input_text, start_id_text, start_id_cmd, start_id_exp

        return input_text, start_id_text, start_id_cmd, start_id_exp

    def ConvertThread(self):
        input_text, start_id_text, start_id_cmd, start_id_exp = self.GetData()
        if len(input_text) == 0 or len(start_id_text) == 0 or len(start_id_cmd) == 0 or len(start_id_exp) == 0:
            return False
        self.id_dic['text_id'] = start_id_text
        self.id_dic['command_id'] = start_id_cmd
        self.id_dic['express_id'] = start_id_exp
        # self.id_dic['text_id'] = IntFormat(int(start_id_text, 16) & 0xFFFFFFFF)
        # self.id_dic['command_id'] = IntFormat(int(start_id_cmd, 16) & 0xFFFFFFFF)
        # self.id_dic['express_id'] = IntFormat(int(start_id_exp, 16) & 0xFFFFFFFF)
        settings.ModifyID(self.id_dic)

        self.ui.lineEdit_4.setText('转换中，请等待....')
        self.ButtonEnabled(False)
        self.convert_thread = ConvertThread(input_text, start_id_text, start_id_cmd, start_id_exp,
                                            self.ui.checkBox_11.isChecked(), self.set_path)
        self.convert_thread.update_date.connect(self.Update)
        self.convert_thread.start()
        # self.convert_thread.wait()
        return True

    def Update(self, text_id, cmd_id, exp_id, old_text, new_text, cmd_text, exp_text):
        self.text_id_str = hex(int(text_id)).replace('0x', '').upper()
        self.cmd_id_str = hex(int(cmd_id)).replace('0x', '').upper()
        self.exp_id_str = hex(int(exp_id)).replace('0x', '').upper()
        self.ui.textEdit_8.setText(old_text)
        self.ui.textEdit_2.setText(new_text)
        self.ui.textEdit_3.setText(cmd_text)
        self.ui.textEdit_4.setText(exp_text)
        self.ui.lineEdit_4.setText('转换完成')
        self.ButtonEnabled(True)

    def Select(self):
        self.set_path = settings.Read()
        if self.set_path:
            path = QFileDialog.getExistingDirectory(self.ui, '请选软件包路径', self.set_path)
        else:
            path = QFileDialog.getExistingDirectory(self.ui, '请选软件包路径', './')
        if path:
            self.set_path = path
            self.ui.lineEdit_2.setText(path)
            settings.Modify(path)

    def SelectExcel(self):
        excel_path = settings.ReadExcelPath()
        if excel_path:
            path = QFileDialog.getExistingDirectory(self.ui, '请选软件包路径', excel_path)
        else:
            path = QFileDialog.getExistingDirectory(self.ui, '请选软件包路径', './')
        if path:
            self.ui.lineEdit_11.setText(path)
            settings.ModifyExcelPath(path)

    @staticmethod
    def AddContent(path, content):
        if content:
            with open(path, 'a+', encoding='gbk') as f:
                f.write(content)

    def Add1(self):
        self.set_path = self.ui.lineEdit_2.text()
        if not self.set_path:
            QMessageBox.warning(self.ui, '错误信息', '软件包文本文件夹路径为空')
            # self.ui.lineEdit_4.setText('软件包文本文件夹路径为空')
            return False
        settings.Modify(self.set_path)
        filename = CheckFileName(self.set_path, 'TEXT')
        if not filename:
            QMessageBox.about(self.ui, '错误信息', 'TEXT文本不存在，请检查软件包路径')
            return False
        elif not self.ui.textEdit_2.toPlainText():
            # QMessageBox.about(self.ui, '错误信息', '无新增TEXT文本ID')
            self.ui.lineEdit_4.setText('无新增TEXT文本ID')
            return True
        else:
            self.AddContent(self.set_path + r'\TEXT_CN.txt', self.ui.textEdit_2.toPlainText())
            self.ui.lineEdit_4.setText('TEXT文本添加完成')
            self.id_dic['text_id'] = self.text_id_str
            self.ui.lineEdit.setText(self.text_id_str)
            settings.ModifyID(self.id_dic)
            return True

    def Add2(self):
        self.set_path = self.ui.lineEdit_2.text()
        if not self.set_path:
            QMessageBox.warning(self.ui, '错误信息', '软件包文本文件夹路径为空')
            # self.ui.lineEdit_4.setText('软件包文本文件夹路径为空')
            return False
        settings.Modify(self.set_path)
        filename = CheckFileName(self.set_path, 'COMMAND')
        if not filename:
            QMessageBox.about(self.ui, '错误信息', 'COMMAND文本不存在，请检查软件包路径')
            return False
        elif not self.ui.textEdit_3.toPlainText():
            # QMessageBox.about(self.ui, '错误信息', '无新增COMMAND文本ID')
            self.ui.lineEdit_4.setText('无新增COMMAND文本ID')
            return True
        else:
            self.AddContent(self.set_path + r'\command.txt', self.ui.textEdit_3.toPlainText())
            self.ui.lineEdit_4.setText('COMMAND文本添加完成')
            self.id_dic['command_id'] = self.cmd_id_str
            self.ui.lineEdit_5.setText(self.cmd_id_str)
            settings.ModifyID(self.id_dic)
            return True

    def Add3(self):
        self.set_path = self.ui.lineEdit_2.text()
        if not self.set_path:
            QMessageBox.warning(self.ui, '错误信息', '软件包文本文件夹路径为空')
            # self.ui.lineEdit_4.setText('软件包文本文件夹路径为空')
            return False
        settings.Modify(self.set_path)
        filename = CheckFileName(self.set_path, 'EXPRESS')
        if not filename:
            QMessageBox.about(self.ui, '错误信息', 'EXPRESS文本不存在，请检查软件包路径')
            return False
        elif not self.ui.textEdit_4.toPlainText():
            # QMessageBox.about(self.ui, '错误信息', '无新增EXPRESS文本ID')
            self.ui.lineEdit_4.setText('无新增EXPRESS文本ID')
            return True
        else:
            self.AddContent(self.set_path + r'\express.txt', self.ui.textEdit_4.toPlainText())
            self.ui.lineEdit_4.setText('EXPRESS文本添加完成')
            self.id_dic['express_id'] = self.exp_id_str
            self.ui.lineEdit_6.setText(self.exp_id_str)
            settings.ModifyID(self.id_dic)
            return True

    def AddAll(self):
        # flag = self.Convert()
        # flag = self.ConvertThread()
        # self.convert_thread.wait()
        # if not flag:
        #     return

        # 一键添加暂时做不到先转换再添加 2023-7-8
        # input_text, start_id_text, start_id_cmd, start_id_exp = self.GetData()
        #
        # self.ui.lineEdit_4.setText('转换中，请等待....')
        # self.ButtonEnabled(False)
        # self.convert_thread = ConvertThread(input_text, start_id_text, start_id_cmd, start_id_exp,
        #                                     self.ui.checkBox_11.isChecked(), self.set_path)
        # self.convert_thread.update_date.connect(self.Update)
        # self.convert_thread.start()
        # self.convert_thread.wait()

        if not self.Add1():  # TEXT
            return
        if not self.Add2():  # COMMAND
            return
        if not self.Add3():  # EXPRESS
            return
        self.ui.lineEdit_4.setText('一键添加完成')

    def RefreshTxtList(self):
        self.ui.comboBox.clear()
        self.set_path = self.ui.lineEdit_2.text()
        if not self.set_path or not os.path.isdir(self.set_path):
            self.ui.lineEdit_4.setText('软件包文本文件夹路径为空')
            return False
        temp_list = []
        txt_list = os.listdir(self.set_path)
        for i in txt_list:
            if os.path.splitext(i)[1] == '.txt' or os.path.splitext(i)[1] == '.TXT':
                temp_list.append(i)
        if not temp_list:
            self.ui.lineEdit_4.setText('选择的文件夹中无txt文件')
            return False
        if not self.ui.checkBox.isChecked() and not self.ui.checkBox_2.isChecked():
            show_txt_list = []
        elif self.ui.checkBox.isChecked() and not self.ui.checkBox_2.isChecked():
            # show_txt_list = [name for name in temp_list if '_EN' not in name]
            show_txt_list = [name for name in temp_list if re.findall(r'_cn', name, re.I) or '_' not in name]
        elif not self.ui.checkBox.isChecked() and self.ui.checkBox_2.isChecked():
            # show_txt_list = [name for name in temp_list if '_CN' not in name]
            show_txt_list = [name for name in temp_list if re.findall(r'_en', name, re.I) or '_' not in name]
        else:
            show_txt_list = temp_list
        # print(show_txt_list)

        if len(show_txt_list) == 0:
            self.ui.comboBox.clear()
        self.ui.comboBox.adjustSize()
        for txt in show_txt_list:
            self.ui.comboBox.addItem(txt)
        if self.first:
            if 'TEXT_CN.txt' in show_txt_list:
                index = show_txt_list.index('TEXT_CN.txt')
                self.ui.comboBox.setCurrentIndex(index)
                self.first = False
            if 'TEXT_CN.TXT' in show_txt_list:
                index = show_txt_list.index('TEXT_CN.TXT')
                self.ui.comboBox.setCurrentIndex(index)
                self.first = False
        return True

    def GetTCodeList(self):
        self.set_path = self.ui.lineEdit_2.text()
        if not self.set_path:
            self.ui.lineEdit_4.setText('软件包文本文件夹路径为空')
            return False
        temp_list = []
        dtc_file_list = []
        txt_list = os.listdir(self.set_path)
        for i in txt_list:
            if os.path.splitext(i)[1] == '.txt' or os.path.splitext(i)[1] == '.TXT':
                temp_list.append(i)

        # 筛选code文本
        if self.ui.checkBox_5.isChecked():
            dtc_file_list = [name for name in temp_list if re.match(r'tcode\d*_', name, re.I)]
            if not dtc_file_list:
                return []
        # 筛选语言
        if not self.ui.checkBox.isChecked() and not self.ui.checkBox_2.isChecked():
            code_list = []
        elif self.ui.checkBox.isChecked() and not self.ui.checkBox_2.isChecked():
            code_list = [name for name in dtc_file_list if re.findall(r'_cn', name, re.I)]
        elif not self.ui.checkBox.isChecked() and self.ui.checkBox_2.isChecked():
            code_list = [name for name in dtc_file_list if re.findall(r'_en', name, re.I)]
        else:
            code_list = dtc_file_list

        # print(code_list)
        return code_list

    @staticmethod
    def isID(org_str):
        result = re.findall(',0x', org_str)
        result1 = re.findall(',0X', org_str)
        if result or result1:
            return True
        else:
            return False

    def hex_format(self, org_str):
        org_str = org_str.replace(' ', '').replace('，', ',')
        result = re.findall(',0x', org_str)
        result1 = re.findall(',0X', org_str)
        if self.isID(org_str):
            id_str = org_str
        else:
            pattern = re.compile('.{2}')
            temp_str = org_str.replace('0X', '').replace('0x', '')
            lists = pattern.findall(temp_str)
            if len(org_str) % 2:
                last_id = '0' + org_str[-1]
                lists.append(last_id)
            id_str = '0x' + ',0x'.join(lists)
        return id_str

    def MatchAndShow(self, index, search_content, option_text):
        items = []
        code = False
        if re.match(r'tcode\d*_', option_text, re.I):  # tcodexx_
            code = True

        if search_content:
            path = self.set_path + '/' + option_text
            f = open(path, 'r', encoding='gbk')
            while True:
                search_flag = False
                line_text = f.readline()
                if not line_text:  # 文件末尾
                    break
                if not line_text.strip():  # 空白行
                    continue
                if line_text[0] == ';':  # 注释行
                    continue

                if code:  # tcodexx_
                    pos = line_text.find('\"')
                    if pos == -1:
                        continue
                    if line_text.rfind(';;') != -1:
                        last_pos = line_text.rindex(';;')
                        text_list = [line_text[0:pos], line_text[pos:last_pos]]
                    else:
                        if line_text.rfind(';') != -1:
                            last_pos = line_text.rindex(';')
                            text_list = [line_text[0:pos].strip(), line_text[pos:last_pos].strip()]
                        else:
                            text_list = [line_text[0:pos].strip(), line_text[pos:].strip()]
                else:
                    text_list = line_text.split('\"')
                    if len(text_list) < 2:  # command文本
                        cmd_list = line_text.split('\t')  # 存在多段命令
                        cmd_str = ''
                        for i in range(1, len(cmd_list)):
                            if cmd_list[i].rfind(';;') != -1:
                                last_pos = cmd_list[i].rindex(';;')
                                cmd_str += cmd_list[i][0:last_pos] + '\t'
                            else:
                                cmd_str += cmd_list[i] + '\t'
                        text_list = [cmd_list[0].strip(), cmd_str.strip()]
                    else:  # text,express...
                        start_pos = line_text.find('\"')
                        if start_pos == -1:
                            continue
                        comment_pos = line_text.rfind(';;')
                        if comment_pos == -1:
                            comment_pos = line_text.rfind(';')
                        end_pos = line_text.rfind('\"')

                        # 不考虑注释中有"的情况
                        text_list = [line_text[0:start_pos].strip(), line_text[start_pos + 1:end_pos].strip()]
                        '''if comment_pos != -1:
                            if end_pos < comment_pos:
                                text_list = [line_text[0:start_pos].strip(), line_text[start_pos+1:end_pos].strip()]
                            else:
                                text_list = [line_text[0:start_pos].strip(), line_text[start_pos+1:comment_pos].strip()]
                        else:
                            text_list = [line_text[0:start_pos].strip(), line_text[start_pos+1:end_pos].strip()]'''

                if index == 0:  # 匹配ID，先把ID格式化
                    search_content = self.hex_format(search_content)
                    # print(search_content)
                text_part = text_list[index]
                # print(text_part)
                if self.ui.checkBox_3.isChecked():  # 全字匹配
                    if self.ui.checkBox_4.isChecked():  # 匹配大小写
                        # if re.fullmatch(search_content, text_part):  # 有缺陷，无法匹配search_content中存在需要转义的字符
                        if search_content == text_part:
                            search_flag = True
                    else:
                        # if re.fullmatch(search_content, text_part, re.I):
                        if search_content.lower() == text_part.lower():
                            search_flag = True
                else:
                    # if text_part.find(search_content) != -1:
                    #     search_flag = True
                    escaped_pattern = re.escape(search_content)  # 防止转义字符 2024-3-9

                    if self.ui.checkBox_4.isChecked():  # 匹配大小写
                        if re.findall(escaped_pattern, text_part):
                            search_flag = True
                    else:
                        if re.findall(escaped_pattern, text_part, re.I):
                            search_flag = True

                if search_flag:
                    item = [text_list[0], text_list[1]]
                    items.append(item)
                    self.line += 1
                    if self.line >= self.MAX_LINE:
                        break
            f.close()
        else:
            self.ui.lineEdit_4.setText('查找内容为空，请输入')
            return

        # print(items)
        if len(items) == 0:
            self.ui.lineEdit_4.setText(f'{option_text}无法找到文本"{search_content}"')
        else:
            # num = 0
            for data in items:
                item11 = QStandardItem(data[0].strip())
                item12 = QStandardItem(data[1].strip())
                item13 = QStandardItem(option_text)
                self.model.setItem(self.num, 0, item11)
                self.model.setItem(self.num, 1, item12)
                self.model.setItem(self.num, 2, item13)
                self.num += 1
            # table_widget = self.ui.tableView
            # table_width = table_widget.width()
            # col_1_width = table_width * 1 // 10
            # col_2_width = table_width * 8 // 10
            # col_3_width = table_width * 1 // 10
            # print(table_widget.size(), table_width, col_1_width, col_2_width, col_3_width)
            # table_widget.setColumnWidth(0, col_1_width)
            # table_widget.setColumnWidth(1, col_2_width)
            # table_widget.setColumnWidth(2, col_3_width)
            self.ui.tableView.setModel(self.model)
            # self.ui.lineEdit_4.setStyleSheet("color:green")
            self.ui.lineEdit_4.setText(f'已找到文本"{search_content}"\t计数:{len(items)}次')
            # self.ui.lineEdit_4.setStyleSheet("color:red")
        return len(items)

    def Search(self, index):
        self.line = 0
        self.num = 0
        self.InitTableView()
        self.set_path = self.ui.lineEdit_2.text()
        if not self.set_path:
            self.ui.lineEdit_4.setText('软件包文本文件夹路径为空')
            return False
        settings.Modify(self.set_path)
        search_content = self.ui.lineEdit_3.text()
        option_text = self.ui.comboBox.currentText()

        # 08 FC 00 02 10 01 00 00 00 00 00 -> 0x08,0xFC,0x00,0x02,0x10,0x01,0x00,0x00,0x00,0x00,0x00查找
        if re.match('command', option_text, re.I):
            search_content = search_content.replace(' ', ',0x')

        if self.ui.checkBox_5.isChecked():
            tcode_files = self.GetTCodeList()
            if tcode_files:
                self.count = 0
                for option_text in tcode_files:
                    # print(option_text)
                    self.count += self.MatchAndShow(index, search_content, option_text)
                self.ui.lineEdit_4.setText(f'"{search_content}"全文搜索故障码完成\t计数:{self.count}次')
            else:
                self.ui.lineEdit_4.setText('无故障码文本')
                return
        else:
            self.MatchAndShow(index, search_content, option_text)

    def CopyIdData(self):
        # print(self.ui.tableView.selectedIndexes())
        data = self.ui.tableView.currentIndex().data()
        if not data:
            self.ui.lineEdit_4.setText('未选中有数据的单元格')
            return
        copy_data = data.replace(',0x', '').replace(',0X', '').replace(',', '')
        copy_data = copy_data.strip()
        # print(copy_data)
        self.ui.lineEdit_4.setText('复制完成')
        clipboard = QApplication.clipboard()
        clipboard.setText(copy_data)

    def CopyAnyData(self):
        # 获取QTableView的选中模型
        selection_model = self.ui.tableView.selectionModel()

        # 获取选中的索引
        selected_indexes = selection_model.selectedIndexes()
        # print(selected_indexes)
        if not selected_indexes:
            self.ui.lineEdit_4.setText('未选中有数据的单元格')
            return

        # 获取选中单元格的值
        selected_values = []
        for index in selected_indexes:
            index_data = index.data()
            if self.isID(index_data):  # 如果是ID，去掉末尾,号
                if index_data[-1] == ',':
                    index_data = index_data[:-1]
            if index_data:
                selected_values.append(index_data)

        # 打印选中单元格的值
        # print(selected_values)
        if len(selected_values) >= 2:
            data = '\n'.join(selected_values)
        else:
            data = ''.join(selected_values).strip()
        # print(data)
        self.ui.lineEdit_4.setText('复制完成')
        clipboard = QApplication.clipboard()
        clipboard.setText(data)

    @staticmethod
    def GetExcelNameList(path, suffix):
        input_template_All = []
        input_template_All_Path = []
        for root, dirs, files in os.walk(path, topdown=False):
            for name in files:
                # print(os.path.join(root, name))
                # print(name)
                if os.path.splitext(name)[1] == suffix:
                    if '~$' not in name:  # ~$ 正打开的Excel文件
                        input_template_All.append(name)
                        input_template_All_Path.append(os.path.join(root, name))
        return input_template_All_Path

    def ReadExcelGetText(self):
        self.ui.plainTextEdit.clear()
        excel_path = self.ui.lineEdit_11.text()
        if not excel_path:
            QMessageBox.warning(self.ui, '错误信息', '未选择Excel协议文档')
            return False
        settings.ModifyExcelPath(excel_path)

        excel_path_list = []
        if os.path.isfile(excel_path):
            if os.path.splitext(os.path.basename(excel_path))[1] == '.xlsx':
                excel_path_list = [excel_path]
        else:
            excel_path_list = self.GetExcelNameList(excel_path, '.xlsx')

        # print(excel_path_list)
        if len(excel_path_list) == 0:
            QMessageBox.about(self.ui, '错误信息', '不存在.xlsx后缀名的协议文件')
            QMessageBox.about(self.ui, '错误信息', '不存在.xlsx后缀名的协议文件')
            return False

        self.ui.lineEdit_4.setText('提取Excel协议文本中，请等待....')
        # input_text = ''
        for excel_path in excel_path_list:
            workbook = openpyxl.load_workbook(excel_path)
            for sheet_name in workbook.sheetnames:
                sheet_text_list = []
                if re.fullmatch('特殊功能菜单', sheet_name.strip()):
                    sheet = workbook[sheet_name]
                    rows = sheet.max_row  # 行数
                    column = sheet.max_column  # 列数
                    # 按列获取值
                    # for xx in sheet.iter_cols(min_row=2, max_row=5, min_col=1, max_col=2):
                    # 按行获取值
                    for i in sheet.iter_rows(min_row=2, max_row=rows, min_col=1, max_col=column):
                        for j in i:
                            if j.value:
                                sheet_text_list.append(j.value)
                                # input_text += j.value + '\n'
                    self.ui.plainTextEdit.insertPlainText('\n'.join(sheet_text_list) + '\n')
                if re.fullmatch('特殊功能', sheet_name.strip()):
                    sheet = workbook[sheet_name]
                    rows = sheet.max_row  # 行数
                    column = sheet.max_column  # 列数
                    spec_func_head = False
                    spec_func_step = True
                    for i in sheet.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=3):
                        num = 1
                        if spec_func_head:
                            spec_func_head = False
                            spec_func_step = True
                            continue
                        for j in i:
                            if j.value:
                                # print(j.value)
                                if re.match('Menu:', str(j.value)):
                                    spec_func_head = True
                                    spec_func_step = False
                                    break
                                if spec_func_step and num < 2:
                                    num += 1
                                    continue
                                cell_value = self.parseData(str(j.value))
                                if len(cell_value):
                                    sheet_text_list.append('\n'.join(cell_value))
                                # input_text += str(j.value).replace('\n', '') + '\n'
                            else:
                                num += 1
                    self.ui.plainTextEdit.insertPlainText('\n'.join(sheet_text_list) + '\n')
            workbook.close()
        self.ui.lineEdit_4.setText('提取文本完成')

    def replace_line_feed(self):
        # 只从剪切板取值
        clipboard = QApplication.clipboard()
        tag_str = clipboard.text()

        if tag_str:
            out_str = tag_str.replace('\n', r'\n')
            clipboard.setText(out_str)
            self.ui.lineEdit_4.setText('替换完成: ' + out_str)
        else:
            self.ui.lineEdit_4.setText('剪切板无内容')

    @staticmethod
    def FormatID(org_str):
        org_str = org_str.replace(' ', '').replace('，', ',')
        result = re.findall(',0x', org_str)
        result1 = re.findall(',0X', org_str)
        if result or result1:
            id_str = org_str.replace(',0x', '').replace(',0X', '')
        else:
            pattern = re.compile('.{2}')
            temp_str = org_str.replace('0X', '').replace('0x', '').upper()
            lists = pattern.findall(temp_str)
            if len(org_str) % 2:
                last_id = '0' + org_str[-1]
                lists.append(last_id)
            id_str = '0x' + ',0x'.join(lists)
        return id_str

    def id_format_show(self):
        clipboard = QApplication.clipboard()
        tag_str = clipboard.text()

        if tag_str:
            id_str = self.FormatID(tag_str)
            clipboard.setText(id_str)
            self.ui.lineEdit_4.setText('ID转换完成: ' + id_str)
        else:
            self.ui.lineEdit_4.setText('剪切板无内容')


if __name__ == '__main__':
    settings.Create()
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon('ui/KO.png'))
    stats = AddID()
    stats.ui.show()
    app.exec_()
